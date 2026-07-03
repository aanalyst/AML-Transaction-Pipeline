import psycopg2
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv
import os

load_dotenv()

# ── 1. Connect to PostgreSQL ──────────────────────────────────────────────────
conn = psycopg2.connect(
    host=os.getenv("DB_HOST"),
    port=os.getenv("DB_PORT"),
    database=os.getenv("DB_NAME"),
    user=os.getenv("DB_USER"),
    password=os.getenv("DB_PASSWORD")
)
cursor = conn.cursor()

# ── 2. Fetch flagged transactions with narratives ─────────────────────────────
cursor.execute("""
    SELECT transaction_id, customer_id, amount, date, country, 
           risk_score, risk_band, sar_narrative
    FROM transactions
    WHERE is_flagged = TRUE
    ORDER BY risk_score DESC
""")
rows = cursor.fetchall()
print(f"Fetched {len(rows)} flagged transactions")

cursor.close()
conn.close()

# ── 3. Create workbook ────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "SAR Report"

# ── 4. Title row ──────────────────────────────────────────────────────────────
ws.merge_cells("A1:H1")
title_cell = ws["A1"]
title_cell.value = "AML Suspicious Activity Report — AUSTRAC Compliance"
title_cell.font = Font(bold=True, size=14)
title_cell.alignment = Alignment(horizontal="center")

# ── 5. Header row ─────────────────────────────────────────────────────────────
headers = ["Transaction ID", "Customer ID", "Amount", "Date", 
           "Country", "Risk Score", "Risk Band", "SAR Narrative"]

header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

for col_num, header in enumerate(headers, start=1):
    cell = ws.cell(row=2, column=col_num, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

# ── 6. Freeze header row ──────────────────────────────────────────────────────
ws.freeze_panes = "A3"

# ── 7. Write data rows ────────────────────────────────────────────────────────
risk_colors = {
    "High": "FF0000",
    "Medium": "FFA500",
    "Low": "00B050"
}

for row_num, row in enumerate(rows, start=3):
    transaction_id, customer_id, amount, date, country, risk_score, risk_band, narrative = row

    ws.cell(row=row_num, column=1, value=transaction_id)
    ws.cell(row=row_num, column=2, value=customer_id)
    ws.cell(row=row_num, column=3, value=float(amount)).number_format = '"$"#,##0.00'
    ws.cell(row=row_num, column=4, value=str(date))
    ws.cell(row=row_num, column=5, value=country)
    ws.cell(row=row_num, column=6, value=round(float(risk_score), 4))
    
    risk_cell = ws.cell(row=row_num, column=7, value=risk_band)
    risk_cell.font = Font(bold=True, color=risk_colors.get(risk_band, "000000"))
    
    ws.cell(row=row_num, column=8, value=narrative).alignment = Alignment(wrap_text=True)

# ── 8. Column widths ──────────────────────────────────────────────────────────
column_widths = [18, 14, 14, 22, 18, 12, 12, 60]
for col_num, width in enumerate(column_widths, start=1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

# ── 9. Save ───────────────────────────────────────────────────────────────────
output_path = "output/sar_report.xlsx"
wb.save(output_path)
print(f"SAR report saved to {output_path}")